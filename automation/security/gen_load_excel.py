#!/usr/bin/env python3
"""
Generate Excel report for load test results.
Reads k6 metrics from k6-metrics.json (handleSummary export)
and pytest results from loadtest_results.json.
Shows all 400 test cases in the results sheet.
"""
import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

os.makedirs("LoadTest_Results/excel", exist_ok=True)

wb = openpyxl.Workbook()

# ─── Colour palette ────────────────────────────────────────────
BG_DARK   = "0d1117"
BG_CARD   = "161b22"
BG_HEADER = "1e1e2e"
C_PURPLE  = "a855f7"
C_GREEN   = "238636"
C_RED     = "da3633"
C_WHITE   = "ffffff"
C_GRAY    = "8b949e"
C_YELLOW  = "e2b714"
C_CYAN    = "22d3ee"

def hdr_fill(color=BG_HEADER):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def hdr_font(color=C_PURPLE, bold=True, size=11):
    return Font(color=color, bold=bold, size=size)

def cell_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_border():
    s = Side(style="thin", color="30363d")
    return Border(left=s, right=s, top=s, bottom=s)

# ══════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = C_PURPLE
ws_dash.column_dimensions["A"].width = 35
ws_dash.column_dimensions["B"].width = 28

def add_header(ws, title, row_offset=0):
    row = ws.max_row + 1 + row_offset
    ws.append([title, ""])
    for cell in ws[ws.max_row]:
        cell.fill = hdr_fill(BG_HEADER)
        cell.font = hdr_font(C_PURPLE, bold=True, size=12)

def add_row(ws, label, value, val_color=C_WHITE):
    ws.append([label, value])
    row = ws.max_row
    ws.cell(row, 1).fill = hdr_fill(BG_CARD)
    ws.cell(row, 1).font = Font(color=C_GRAY, size=11)
    ws.cell(row, 2).fill = hdr_fill(BG_CARD)
    ws.cell(row, 2).font = Font(color=val_color, bold=True, size=11)

# Title
ws_dash.append(["FlowPay Load and Performance Test Report"])
ws_dash["A1"].font = Font(bold=True, size=18, color=C_PURPLE)
ws_dash["A1"].fill = hdr_fill(BG_DARK)
ws_dash.append(["Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws_dash["A2"].font = Font(color=C_GRAY, size=11)
ws_dash.append(["Target URL: " + os.environ.get("BASE_URL", "https://jananibsse.github.io/janani_pdd_flowpay/")])
ws_dash["A3"].font = Font(color=C_GRAY, size=11)
ws_dash.append([])

# ── k6 Load Test Metrics ──────────────────────────────────────
# Try new format first (k6-metrics.json from handleSummary)
k6_data = {}
k6_files = [
    "LoadTest_Results/k6/k6-metrics.json",
    "LoadTest_Results/k6/baseline-summary.json",
]
for k6f in k6_files:
    if os.path.exists(k6f):
        with open(k6f) as fh:
            raw = json.load(fh)
        # Detect format
        if "metrics" in raw and "response_time" in raw.get("metrics", {}):
            # New handleSummary format
            k6_data = raw["metrics"]
        elif "metrics" in raw:
            # Old summary-export format — drill into .values
            m = raw["metrics"]
            d_vals = m.get("http_req_duration", {})
            r_vals = m.get("http_reqs", {})
            f_vals = m.get("http_req_failed", {})
            # summary-export stores values directly at top level of each metric
            k6_data = {
                "rps": r_vals.get("rate", r_vals.get("values", {}).get("rate", 0)),
                "total_requests": r_vals.get("count", r_vals.get("values", {}).get("count", 0)),
                "response_time": {
                    "avg_ms": d_vals.get("avg", d_vals.get("values", {}).get("avg", None)),
                    "min_ms": d_vals.get("min", d_vals.get("values", {}).get("min", None)),
                    "med_ms": d_vals.get("med", d_vals.get("values", {}).get("med", None)),
                    "max_ms": d_vals.get("max", d_vals.get("values", {}).get("max", None)),
                    "p90_ms": d_vals.get("p(90)", d_vals.get("values", {}).get("p(90)", None)),
                    "p95_ms": d_vals.get("p(95)", d_vals.get("values", {}).get("p(95)", None)),
                    "p99_ms": d_vals.get("p(99)", d_vals.get("values", {}).get("p(99)", None)),
                },
                "error_rate_pct": (
                    f_vals.get("rate", f_vals.get("values", {}).get("rate", 0))
                ) * 100,
            }
        break

def fmt_ms(val):
    if val is None or val == 0:
        return "N/A"
    return str(round(val, 1)) + " ms"

def fmt_val(val, suffix=""):
    if val is None:
        return "N/A"
    return str(round(val, 2)) + suffix

rt = k6_data.get("response_time", {})
rps = k6_data.get("rps", None)
total_reqs = k6_data.get("total_requests", 0)
err_pct = k6_data.get("error_rate_pct", None)

add_header(ws_dash, "k6 Baseline Load Test Metrics (100 VUs x 1 min)")
add_row(ws_dash, "Requests Per Second", fmt_val(rps, " req/sec"), C_CYAN)
add_row(ws_dash, "Total HTTP Requests", str(int(total_reqs)) if total_reqs else "N/A", C_WHITE)
add_row(ws_dash, "", "", C_WHITE)
add_header(ws_dash, "Response Times (End-to-End, including network)")
add_row(ws_dash, "Average Response Time", fmt_ms(rt.get("avg_ms")), C_CYAN)
add_row(ws_dash, "Median (P50) Response", fmt_ms(rt.get("med_ms")), C_WHITE)
add_row(ws_dash, "Minimum Response Time", fmt_ms(rt.get("min_ms")), C_GRAY)
add_row(ws_dash, "Maximum Response Time", fmt_ms(rt.get("max_ms")), C_WHITE)
add_row(ws_dash, "P90 Response Time", fmt_ms(rt.get("p90_ms")), C_WHITE)
add_row(ws_dash, "P95 Response Time", fmt_ms(rt.get("p95_ms")), C_YELLOW)
add_row(ws_dash, "P99 Response Time", fmt_ms(rt.get("p99_ms")), C_YELLOW)
add_row(ws_dash, "Error Rate", fmt_val(err_pct, "%"), C_GREEN if (err_pct or 0) < 5 else C_RED)

min_note = "Note: Min reflects CDN HTTP/2 connection reuse / cache hit. Avg/P95/P99 are more representative."
ws_dash.append([min_note])
ws_dash[ws_dash.max_row][0].font = Font(color=C_GRAY, italic=True, size=10)
ws_dash.append([])

# ── pytest Assertion Summary ──────────────────────────────────
rf = "LoadTest_Results/json/loadtest_results.json"
all_tests = []
passed = failed = 0
if os.path.exists(rf):
    with open(rf) as fh:
        data = json.load(fh)
    for t in data.get("tests", []):
        name = t.get("nodeid", "")
        outcome = t.get("outcome", "")
        dur = t.get("duration", 0)
        all_tests.append((name, outcome, dur))
        if outcome == "passed":
            passed += 1
        else:
            failed += 1

total_assertions = passed + failed
rate = (passed / total_assertions * 100) if total_assertions > 0 else 0

add_header(ws_dash, "400 Assertion Results (pytest)")
add_row(ws_dash, "Total Assertions", str(total_assertions), C_WHITE)
add_row(ws_dash, "Passed", str(passed), C_GREEN)
add_row(ws_dash, "Failed", str(failed), C_RED if failed > 0 else C_WHITE)
add_row(ws_dash, "Pass Rate", str(round(rate, 1)) + "%",
        C_GREEN if rate >= 90 else C_YELLOW if rate >= 70 else C_RED)

# ── Category Breakdown ────────────────────────────────────────
ws_dash.append([])
add_header(ws_dash, "Test Category Breakdown")
categories_expected = {
    "Response Time Validation":  50,
    "Throughput Analysis":       50,
    "Error Rate Validation":     50,
    "Concurrency Testing":       50,
    "Endpoint Performance":      50,
    "Resource and Efficiency":   50,
    "Scalability Checks":        50,
    "Reliability and Stability": 50,
}

def get_category(name):
    n = name.lower()
    if "TestResponseTime" in name or "test_rt_" in name:
        return "Response Time Validation"
    if "TestThroughput" in name or "test_tp_" in name:
        return "Throughput Analysis"
    if "TestErrorRate" in name or "test_er_" in name:
        return "Error Rate Validation"
    if "TestConcurren" in name or "test_cc_" in name:
        return "Concurrency Testing"
    if "TestEndpoint" in name or "test_ep_" in name:
        return "Endpoint Performance"
    if "TestResource" in name or "test_res_" in name:
        return "Resource and Efficiency"
    if "TestScalabil" in name or "test_sc_" in name:
        return "Scalability Checks"
    if "TestReliabil" in name or "test_rel_" in name:
        return "Reliability and Stability"
    return "General"

cat_stats = {}
for name, outcome, dur in all_tests:
    cat = get_category(name)
    if cat not in cat_stats:
        cat_stats[cat] = {"passed": 0, "failed": 0}
    if outcome == "passed":
        cat_stats[cat]["passed"] += 1
    else:
        cat_stats[cat]["failed"] += 1

ws_dash.append(["Category", "Expected", "Passed", "Failed", "Pass Rate"])
for cell in ws_dash[ws_dash.max_row]:
    cell.fill = hdr_fill(BG_HEADER)
    cell.font = hdr_font(C_PURPLE)

for cat, exp in categories_expected.items():
    s = cat_stats.get(cat, {"passed": 0, "failed": 0})
    cat_total = s["passed"] + s["failed"]
    cat_rate = (s["passed"] / cat_total * 100) if cat_total > 0 else 0
    ws_dash.append([cat, exp, s["passed"], s["failed"], str(round(cat_rate, 0)) + "%"])
    row = ws_dash.max_row
    for c in range(1, 6):
        ws_dash.cell(row, c).fill = hdr_fill(BG_CARD)
        ws_dash.cell(row, c).font = Font(color=C_GRAY, size=11)
    ws_dash.cell(row, 5).font = Font(
        color=C_GREEN if cat_rate >= 90 else C_YELLOW if cat_rate >= 70 else C_RED,
        bold=True, size=11
    )

for col in ["A","B","C","D","E","F"]:
    ws_dash.column_dimensions[col].width = 30

# ══════════════════════════════════════════════════════════════
# SHEET 2 — ALL 400 TEST RESULTS
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("400 Test Results")
ws2.sheet_properties.tabColor = C_GREEN

headers = ["#", "Test Case ID", "Test Name", "Category", "Status", "Duration (s)", "Response Time Context"]
ws2.append(headers)
for cell in ws2[1]:
    cell.fill = hdr_fill(BG_HEADER)
    cell.font = hdr_font(C_PURPLE, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center")

pass_fill = cell_fill(C_GREEN)
fail_fill = cell_fill(C_RED)
pass_font = Font(color=C_WHITE, bold=True, size=10)
fail_font = Font(color=C_WHITE, bold=True, size=10)

for i, (name, outcome, dur) in enumerate(all_tests, 1):
    parts = name.split("::")
    test_id = parts[-1] if len(parts) > 1 else name
    # Extract friendly name from test ID (remove prefix like test_rt_001_)
    friendly = test_id
    if "_" in test_id and len(test_id.split("_")) > 3:
        friendly = " ".join(test_id.split("_")[3:]).replace("_", " ").title()
    cat = get_category(name)
    status = outcome.upper()

    # Response time context note
    rt_note = ""
    if "response" in test_id or "rt_" in test_id:
        rt_note = "Measured via requests.get() — includes DNS + TCP + TLS + server"
    elif "throughput" in test_id or "tp_" in test_id:
        rt_note = "Sequential HTTP requests — throughput measurement"
    elif "concurrent" in test_id or "cc_" in test_id:
        rt_note = "Multi-threaded concurrent request batch"
    else:
        rt_note = "HTTP assertion test"

    ws2.append([i, test_id, friendly, cat, status, str(round(dur, 3)), rt_note])
    row = ws2.max_row
    for c in range(1, 8):
        ws2.cell(row, c).fill = hdr_fill(BG_CARD)
        ws2.cell(row, c).font = Font(color=C_GRAY, size=10)

    sc = ws2.cell(row, 5)
    if outcome == "passed":
        sc.fill = pass_fill
        sc.font = pass_font
    else:
        sc.fill = fail_fill
        sc.font = fail_font

    ws2.cell(row, 1).alignment = Alignment(horizontal="center")
    ws2.cell(row, 5).alignment = Alignment(horizontal="center")

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 28
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 50

ws2.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════
# SHEET 3 — k6 METRICS DETAIL
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("k6 Metrics Detail")
ws3.sheet_properties.tabColor = C_CYAN

ws3.append(["FlowPay k6 Load Test — Detailed Metrics"])
ws3["A1"].font = Font(bold=True, size=14, color=C_CYAN)
ws3.append([])
ws3.append(["Metric", "Value", "Threshold", "Status"])
for cell in ws3[3]:
    cell.fill = hdr_fill(BG_HEADER)
    cell.font = hdr_font(C_CYAN)

metrics_rows = [
    ("RPS (Requests/sec)",     fmt_val(rps, ""),          "No threshold",   "INFO"),
    ("Total HTTP Requests",    str(int(total_reqs or 0)), "No threshold",   "INFO"),
    ("Avg Response Time",      fmt_ms(rt.get("avg_ms")),  "< 2000 ms",
        "PASS" if rt.get("avg_ms") and rt["avg_ms"] < 2000 else "REVIEW"),
    ("Median (P50) Response",  fmt_ms(rt.get("med_ms")),  "< 2000 ms",
        "PASS" if rt.get("med_ms") and rt["med_ms"] < 2000 else "REVIEW"),
    ("P90 Response Time",      fmt_ms(rt.get("p90_ms")),  "< 5000 ms",
        "PASS" if rt.get("p90_ms") and rt["p90_ms"] < 5000 else "REVIEW"),
    ("P95 Response Time",      fmt_ms(rt.get("p95_ms")),  "< 5000 ms",
        "PASS" if rt.get("p95_ms") and rt["p95_ms"] < 5000 else "REVIEW"),
    ("P99 Response Time",      fmt_ms(rt.get("p99_ms")),  "< 8000 ms",
        "PASS" if rt.get("p99_ms") and rt["p99_ms"] < 8000 else "REVIEW"),
    ("Min Response Time",      fmt_ms(rt.get("min_ms")),
        "Note: CDN cache hit", "INFO"),
    ("Max Response Time",      fmt_ms(rt.get("max_ms")),  "< 30000 ms",   "INFO"),
    ("Error Rate",             fmt_val(err_pct, "%"),      "< 5%",
        "PASS" if err_pct is not None and err_pct < 5 else "REVIEW"),
]

for m_label, m_val, m_thresh, m_status in metrics_rows:
    ws3.append([m_label, m_val, m_thresh, m_status])
    row = ws3.max_row
    for c in range(1, 5):
        ws3.cell(row, c).fill = hdr_fill(BG_CARD)
        ws3.cell(row, c).font = Font(color=C_GRAY, size=11)
    status_cell = ws3.cell(row, 4)
    if m_status == "PASS":
        status_cell.fill = pass_fill
        status_cell.font = Font(color=C_WHITE, bold=True, size=11)
    elif m_status == "INFO":
        status_cell.font = Font(color=C_CYAN, bold=True, size=11)

ws3.append([])
ws3.append(["Response Time Explanation"])
ws3[ws3.max_row][0].font = Font(bold=True, color=C_YELLOW, size=12)
explanations = [
    "Min Response Time: The absolute fastest response observed. For GitHub Pages CDN, this reflects",
    "  an HTTP/2 multiplexed request on an already-established connection to the nearest CDN edge node.",
    "  While sub-1ms is unusual, it is technically possible when the TCP connection and TLS handshake",
    "  are already complete and the CDN serves the response from memory cache.",
    "",
    "Avg Response Time: The mean across ALL requests. More representative of typical user experience.",
    "",
    "P95/P99 Response Time: 95th/99th percentile — 95%/99% of all requests completed within this time.",
    "  These are the most important SLO metrics for production readiness.",
    "",
    "Recommendation: Focus on Avg, P95, and P99 for SLA compliance. Min is informational only.",
]
for line in explanations:
    ws3.append([line])
    ws3[ws3.max_row][0].font = Font(color=C_GRAY, size=10, italic=(line.startswith("  ")))

for col in ["A","B","C","D"]:
    ws3.column_dimensions[col].width = 35

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
wb.save("LoadTest_Results/excel/LoadTest_Report.xlsx")
print("LoadTest Excel report saved!")
print("  Total assertions: " + str(total_assertions))
print("  Passed: " + str(passed))
print("  Failed: " + str(failed))
print("  Pass rate: " + str(round(rate, 1)) + "%")
print("  Sheets: Dashboard, 400 Test Results, k6 Metrics Detail")
