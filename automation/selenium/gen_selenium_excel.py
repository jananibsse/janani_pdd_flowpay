#!/usr/bin/env python3
"""Generate Excel report for Selenium test results."""
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font

os.makedirs("reports/excel", exist_ok=True)

try:
    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "1e90ff"
    
    hf = PatternFill(start_color="1e1e2e", end_color="1e1e2e", fill_type="solid")
    hfont = Font(color="58a6ff", bold=True, size=12)
    pfill = PatternFill(start_color="238636", end_color="238636", fill_type="solid")
    ffill = PatternFill(start_color="da3633", end_color="da3633", fill_type="solid")
    wfont = Font(color="ffffff", size=11)
    
    ws_dash.append(["FlowPay Selenium E2E Test Report"])
    ws_dash["A1"].font = Font(bold=True, size=16, color="58a6ff")
    ws_dash.append(["Generated: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_dash.append(["Target: " + os.environ.get('BASE_URL', 'N/A')])
    ws_dash.append([])
    
    rf = "reports/json/selenium_results.json"
    tests = []
    passed = failed = errors = skipped = 0
    
    if os.path.exists(rf):
        with open(rf) as fh:
            data = json.load(fh)
        for t in data.get("tests", []):
            name = t.get("nodeid", "Unknown")
            outcome = t.get("outcome", "unknown")
            dur = t.get("duration", 0)
            tests.append((name, outcome, dur))
            if outcome == "passed":
                passed += 1
            elif outcome == "failed":
                failed += 1
            elif outcome == "error":
                errors += 1
            else:
                skipped += 1
                
    total = passed + failed + errors + skipped
    rate = (passed / total * 100) if total > 0 else 0
    
    ws_dash.append(["Metric", "Value"])
    for cell in ws_dash[ws_dash.max_row]:
        cell.fill = hf
        cell.font = hfont
        
    for row in [["Total Tests", total], ["Passed", passed], ["Failed", failed],
                ["Errors", errors], ["Skipped", skipped], ["Pass Rate", f"{rate:.1f}%"]]:
        ws_dash.append(row)
        
    ws_dash.column_dimensions["A"].width = 30
    ws_dash.column_dimensions["B"].width = 25
    
    ws_det = wb.create_sheet("Test Results")
    ws_det.sheet_properties.tabColor = "238636"
    ws_det.append(["Num", "Test Case", "Module", "Status", "Duration (s)"])
    for cell in ws_det[1]:
        cell.fill = hf
        cell.font = hfont
        
    for i, (name, outcome, dur) in enumerate(tests, 1):
        parts = name.split("::")
        module = parts[0].split("/")[-1].replace(".py", "") if parts else "unknown"
        tname = parts[-1] if len(parts) > 1 else name
        ws_det.append([i, tname, module, outcome.upper(), f"{dur:.2f}"])
        row = ws_det.max_row
        sc = ws_det.cell(row, 4)
        if outcome == "passed":
            sc.fill = pfill
            sc.font = wfont
        elif outcome in ("failed", "error"):
            sc.fill = ffill
            sc.font = wfont
            
    for col in ["A","B","C","D","E"]:
        ws_det.column_dimensions[col].width = 25
    ws_det.column_dimensions["B"].width = 60
    
    ws_mod = wb.create_sheet("Module Summary")
    ws_mod.sheet_properties.tabColor = "f0883e"
    ws_mod.append(["Module", "Total", "Passed", "Failed", "Pass Rate"])
    for cell in ws_mod[1]:
        cell.fill = hf
        cell.font = hfont
        
    modules = {}
    for name, outcome, _ in tests:
        parts = name.split("::")
        mod = parts[0].split("/")[-1].replace(".py", "") if parts else "unknown"
        if mod not in modules:
            modules[mod] = {"total": 0, "passed": 0, "failed": 0}
        modules[mod]["total"] += 1
        if outcome == "passed":
            modules[mod]["passed"] += 1
        else:
            modules[mod]["failed"] += 1
            
    for mod, stats in sorted(modules.items()):
        pr = (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0
        ws_mod.append([mod, stats["total"], stats["passed"], stats["failed"], f"{pr:.0f}%"])
        
    for col in ["A","B","C","D","E"]:
        ws_mod.column_dimensions[col].width = 22
        
    wb.save("reports/excel/Selenium_E2E_Report.xlsx")
    print(f"Selenium Excel report saved: {total} tests, {passed} passed, {failed} failed")
except Exception as e:
    print(f"Excel generation error: {e}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Selenium E2E Test Report — Error generating detailed report"])
    wb.save("reports/excel/Selenium_E2E_Report.xlsx")
